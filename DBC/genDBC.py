from typing import Any, Dict, Generator

import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import rows_from_range

TableRow = Dict[str, Any]

def iter_table_rows(ws:Worksheet, tb:Table) -> Generator[TableRow, None, None]:
    """Iterate over rows from a table with headers (row as dictionary)"""
    def get_row_values(row_cell_ref):
        return [ws[c].value for c in row_cell_ref]
    
    iter_rows = rows_from_range(tb.ref)
    headers = get_row_values(next(iter_rows))
    
    for row_cells in iter_rows:
        yield {h:v for h,v in zip(headers, get_row_values(row_cells))}

GEN_LOC = "src/DBC.h"
DBC_LOC = "DBC.xlsx"

START_DOC = """/**
 * AUTO GENERATED FILE - DO NOT EDIT
 * 
 * DBC Values, defined in C as their respective ID
 */

"""
VALUE_LINE = "#define {} {} // {}"

a = openpyxl.load_workbook(DBC_LOC).worksheets[0]

with open(GEN_LOC, "w", encoding="utf-8") as f:
    f.write(START_DOC)
    f.write(VALUE_LINE.format("SPEED_WHEEL_BACK_LEFT", 5, "AMOGUS jajaja"))
